"""
Genera archivos Excel profesionales para reportes diarios y semanales.

Convenciones de diseño:
- Encabezados: fondo azul marino #1B3A6B, texto blanco, negrita
- Filas alternas: blanco puro / azul muy claro #EBF0FA
- Fila de totales / resumen: fondo amarillo #FFF3CD
- Autofilter activado en encabezado
- Anchos de columna ajustados al contenido
- Freeze pane en la primera fila de datos
"""

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta corporativa ────────────────────────────────────────────────────────
_NAVY = "1B3A6B"
_WHITE = "FFFFFF"
_LIGHT_ROW = "EBF0FA"
_SUMMARY_BG = "FFF3CD"
_ALERT_BG = "FDDEDE"
_DELIVERED_BG = "D4EDDA"
_BORDER_COLOR = "C5C5C5"


# ── Estructuras de datos de reporte ──────────────────────────────────────────

@dataclass
class DailyReportRow:
    query_date: date
    query_time: str            # "07:00"
    tracking_number: str
    advisor_name: str
    client_name: str
    current_status: str        # normalizado
    current_status_raw: str    # exacto de TCC
    last_event_at: datetime | None
    hours_without_movement: float | None
    days_without_movement: float | None
    is_delivered: bool
    is_alert: bool
    observations: str
    shipping_date: date | None = None   # fecha de despacho ingresada al registrar
    days_in_transit: int | None = None  # dias desde shipping_date hasta entrega (o hoy)


@dataclass
class WeeklyReportRow:
    week_label: str            # "2026-04-13 al 2026-04-18"
    tracking_number: str
    advisor_name: str
    client_name: str
    first_status: str
    last_status: str
    delivered_at: datetime | None
    total_movements: int
    still_active: bool
    alerts_detected: int
    observations: str


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _header_font() -> Font:
    return Font(name="Calibri", bold=True, color=_WHITE, size=11)


def _body_font(bold: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, size=10)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_row(ws, row_num: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _fill(_NAVY)
        cell.border = _thin_border()
        cell.alignment = _center()


def _apply_data_cell(cell, value, row_idx: int, align=None) -> None:
    cell.value = value
    cell.font = _body_font()
    cell.fill = _fill(_WHITE if row_idx % 2 == 0 else _LIGHT_ROW)
    cell.border = _thin_border()
    cell.alignment = align or _left()


def _set_column_widths(ws, widths: list[int]) -> None:
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_title_block(ws, title: str, subtitle: str) -> None:
    """Inserta dos filas de título al inicio de la hoja."""
    ws.insert_rows(1, amount=3)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=_NAVY)
    title_cell.alignment = _left()

    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Calibri", size=10, color="555555")
    sub_cell.alignment = _left()

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6  # spacer


# ── Servicio principal ────────────────────────────────────────────────────────

class ExcelService:
    """Genera archivos Excel profesionales para reportes TCC."""

    # Columnas del reporte diario
    _DAILY_HEADERS = [
        "Fecha Consulta",
        "Hora Consulta",
        "# Guía",
        "Asesor",
        "Cliente",
        "Estado Actual",
        "Estado Raw (TCC)",
        "Fecha Despacho",
        "Dias en Transito",
        "Última Novedad",
        "Entregada",
        "Alerta 72h",
        "Observaciones",
    ]

    _DAILY_COL_WIDTHS = [14, 12, 18, 22, 25, 18, 35, 16, 16, 22, 12, 12, 40]

    # Columnas del reporte semanal
    _WEEKLY_HEADERS = [
        "Semana",
        "# Guía",
        "Asesor",
        "Cliente",
        "Primer Estado (Semana)",
        "Último Estado (Semana)",
        "Fecha Entrega",
        "Total Movimientos",
        "Siguió Activa",
        "Alertas Detectadas",
        "Observaciones",
    ]

    _WEEKLY_COL_WIDTHS = [26, 18, 22, 25, 24, 24, 20, 18, 14, 18, 40]

    def generate_daily(
        self,
        rows: list[DailyReportRow],
        output_path: Path,
        cycle_label: str,
        report_date: date,
    ) -> Path:
        """Genera el Excel del reporte diario y lo guarda en output_path."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Diario"

        subtitle = (
            f"Generado: {report_date.strftime('%d/%m/%Y')} — Ciclo {cycle_label}  |  "
            f"Total guías: {len(rows)}  |  "
            f"Entregadas: {sum(1 for r in rows if r.is_delivered)}  |  "
            f"Alertas 72h: {sum(1 for r in rows if r.is_alert)}"
        )
        _add_title_block(ws, "ASTECO — Reporte Diario de Guías TCC", subtitle)

        # Encabezado (fila 4 después del título + spacer)
        header_row = 4
        _apply_header_row(ws, header_row, self._DAILY_HEADERS)
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(self._DAILY_HEADERS))}{header_row + len(rows)}"
        )
        ws.freeze_panes = f"A{header_row + 1}"

        # Datos
        for data_idx, row in enumerate(rows, start=1):
            excel_row = header_row + data_idx
            row_bg = _WHITE if data_idx % 2 == 0 else _LIGHT_ROW

            # Override de color por condición especial
            if row.is_delivered:
                row_bg = _DELIVERED_BG
            elif row.is_alert:
                row_bg = _ALERT_BG

            values = [
                row.query_date.strftime("%Y-%m-%d"),
                row.query_time,
                row.tracking_number,
                row.advisor_name,
                row.client_name or "—",
                row.current_status,
                row.current_status_raw or "—",
                row.shipping_date.strftime("%Y-%m-%d") if row.shipping_date else "—",
                row.days_in_transit if row.days_in_transit is not None else "—",
                row.last_event_at.strftime("%Y-%m-%d %H:%M") if row.last_event_at else "—",
                "Sí" if row.is_delivered else "No",
                "Sí" if row.is_alert else "No",
                row.observations or "",
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = _body_font(bold=row.is_delivered or row.is_alert)
                cell.fill = _fill(row_bg)
                cell.border = _thin_border()
                cell.alignment = _center() if col_idx in (1, 2, 8, 9, 11, 12) else _left()

        _set_column_widths(ws, self._DAILY_COL_WIDTHS)
        ws.row_dimensions[header_row].height = 28

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    def generate_weekly(
        self,
        rows: list[WeeklyReportRow],
        week_start: date,
        week_end: date,
        output_path: Path,
    ) -> Path:
        """Genera el Excel del reporte semanal y lo guarda en output_path."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidado Semanal"

        week_label = f"{week_start.strftime('%d/%m/%Y')} al {week_end.strftime('%d/%m/%Y')}"
        subtitle = (
            f"Semana: {week_label}  |  "
            f"Total guías: {len(rows)}  |  "
            f"Entregadas: {sum(1 for r in rows if r.delivered_at)}  |  "
            f"Activas al cierre: {sum(1 for r in rows if r.still_active)}"
        )
        _add_title_block(ws, "ASTECO — Consolidado Semanal de Guías TCC", subtitle)

        header_row = 4
        _apply_header_row(ws, header_row, self._WEEKLY_HEADERS)
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(self._WEEKLY_HEADERS))}{header_row + len(rows)}"
        )
        ws.freeze_panes = f"A{header_row + 1}"

        for data_idx, row in enumerate(rows, start=1):
            excel_row = header_row + data_idx
            row_bg = _WHITE if data_idx % 2 == 0 else _LIGHT_ROW

            if row.delivered_at:
                row_bg = _DELIVERED_BG
            elif row.alerts_detected > 0:
                row_bg = _ALERT_BG

            values = [
                row.week_label,
                row.tracking_number,
                row.advisor_name,
                row.client_name or "—",
                row.first_status or "—",
                row.last_status,
                row.delivered_at.strftime("%Y-%m-%d %H:%M") if row.delivered_at else "—",
                row.total_movements,
                "Sí" if row.still_active else "No",
                row.alerts_detected,
                row.observations or "",
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = _body_font()
                cell.fill = _fill(row_bg)
                cell.border = _thin_border()
                cell.alignment = _center() if col_idx in (7, 8, 9, 10) else _left()

        _set_column_widths(ws, self._WEEKLY_COL_WIDTHS)
        ws.row_dimensions[header_row].height = 28

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path
