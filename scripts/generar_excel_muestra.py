"""
Genera un Excel de muestra con formato definitivo para validación.
Ejecutar: python scripts/generar_excel_muestra.py
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta ─────────────────────────────────────────────────────────────────────
NAVY        = "1B3A6B"
NAVY_LIGHT  = "2C5090"
WHITE       = "FFFFFF"
ROW_LIGHT   = "EBF0FA"
ROW_ALT     = "FFFFFF"
DELIVERED   = "D4EDDA"   # verde claro
ALERT       = "FDE8E8"   # rojo claro
NOVEDAD     = "FFF3CD"   # amarillo claro
TITLE_GRAY  = "F2F4F8"

BORDER_COLOR = "C8D0DA"

# ── Estilos base ───────────────────────────────────────────────────────────────
def _fill(hex6):
    return PatternFill("solid", fgColor=hex6)

def _border(color=BORDER_COLOR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, size=10, color="000000", name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── Datos de muestra ───────────────────────────────────────────────────────────
MUESTRA_DIARIO = [
    # (guia, asesor, cliente, estado_norm, estado_raw, ultima_novedad, horas_sin_mov, entregado, alerta, obs)
    ("TCC-2024-0001", "Juan Pérez",     "Distribuidora Sur SA",    "En Tránsito",   "EN TRÁNSITO - PLANTA BOGOTÁ",            "2026-04-21 08:30",  14.5,  False, False, ""),
    ("TCC-2024-0002", "María López",    "Comercial Norte Ltda",    "En Ruta",       "MENSAJERO EN CAMINO - ENTREGA HOY",       "2026-04-22 06:45",   0.3,  False, False, ""),
    ("TCC-2024-0003", "Carlos Ríos",    "Ferretería Central",      "Entregado",     "ENTREGADO AL DESTINATARIO",               "2026-04-22 09:15",   0.0,  True,  False, "Entregado en este ciclo"),
    ("TCC-2024-0004", "Ana Gómez",      "Supermercados del Valle", "Novedad",       "NOVEDAD - DIRECCIÓN INCORRECTA",          "2026-04-19 14:00",  65.0,  False, False, "Gestionar corrección de dirección"),
    ("TCC-2024-0005", "Pedro Vargas",   "Textiles El Bosque",      "En Tránsito",   "EN TRÁNSITO - SEDE MEDELLÍN",             "2026-04-17 10:00",  93.2,  False, True,  "Sin movimiento 93h — revisar urgente"),
    ("TCC-2024-0006", "Luisa Castro",   "Confecciones Modernas",   "En Tránsito",   "EN TRÁNSITO - PLANTA CALI",               "2026-04-20 16:30",  38.7,  False, False, ""),
    ("TCC-2024-0007", "Jorge Molina",   "Importadora Andina",      "Recogido",      "RECOGIDO POR MENSAJERO - EN PROCESO",     "2026-04-22 07:00",   0.1,  False, False, ""),
    ("TCC-2024-0008", "Sandra Ruiz",    "Electrodomésticos Flash",  "En Tránsito",   "EN TRÁNSITO - HUB BARRANQUILLA",         "2026-04-16 09:00", 118.1,  False, True,  "Sin movimiento 118h — ALERTA CRÍTICA"),
    ("TCC-2024-0009", "Andrés Cano",    "Repuestos San Carlos",    "Entregado",     "ENTREGADO AL DESTINATARIO EN PORTERÍA",   "2026-04-22 08:45",   0.0,  True,  False, "Entregado en este ciclo"),
    ("TCC-2024-0010", "Valentina Cruz", "Papelería El Ángel",      "En Tránsito",   "EN TRÁNSITO - PLANTA BUCARAMANGA",        "2026-04-21 11:00",  20.1,  False, False, ""),
]

MUESTRA_SEMANAL = [
    # (guia, asesor, cliente, primer_estado, ultimo_estado, fecha_entrega, movimientos, activa, alertas, obs)
    ("TCC-2024-0001", "Juan Pérez",     "Distribuidora Sur SA",    "Recogido",  "En Tránsito",   "",                   3,  True,  0, ""),
    ("TCC-2024-0002", "María López",    "Comercial Norte Ltda",    "Recogido",  "En Ruta",       "",                   5,  True,  0, ""),
    ("TCC-2024-0003", "Carlos Ríos",    "Ferretería Central",      "Recogido",  "Entregado",     "2026-04-22 09:15",   6,  False, 0, ""),
    ("TCC-2024-0004", "Ana Gómez",      "Supermercados del Valle", "Recogido",  "Novedad",       "",                   2,  True,  1, "Dirección incorrecta"),
    ("TCC-2024-0005", "Pedro Vargas",   "Textiles El Bosque",      "Recogido",  "En Tránsito",   "",                   2,  True,  1, "Sin movimiento +72h"),
    ("TCC-2024-0006", "Luisa Castro",   "Confecciones Modernas",   "Recogido",  "En Tránsito",   "",                   4,  True,  0, ""),
    ("TCC-2024-0007", "Jorge Molina",   "Importadora Andina",      "Registrado","Recogido",      "",                   1,  True,  0, ""),
    ("TCC-2024-0008", "Sandra Ruiz",    "Electrodomésticos Flash",  "Recogido",  "En Tránsito",   "",                   2,  True,  1, "Sin movimiento +118h"),
    ("TCC-2024-0009", "Andrés Cano",    "Repuestos San Carlos",    "Recogido",  "Entregado",     "2026-04-22 08:45",   5,  False, 0, ""),
    ("TCC-2024-0010", "Valentina Cruz", "Papelería El Ángel",      "Recogido",  "En Tránsito",   "",                   3,  True,  0, ""),
]


def generar_reporte_diario(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "Reporte Diario"

    HOY   = "2026-04-22"
    HORA  = "07:00"
    CICLO = "0700"

    # ── Fila 1: título principal ───────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "ASTECO — Reporte Diario de Guías TCC"
    c.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    c.fill = _fill(TITLE_GRAY)
    c.alignment = _left()
    ws.row_dimensions[1].height = 26

    # ── Fila 2: subtítulo ─────────────────────────────────────────────────────
    ws.merge_cells("A2:M2")
    c = ws["A2"]
    entregadas = sum(1 for r in MUESTRA_DIARIO if r[7])
    alertas    = sum(1 for r in MUESTRA_DIARIO if r[8])
    c.value = (
        f"Fecha: {HOY}   |   Ciclo: {HORA}   |   "
        f"Total guías: {len(MUESTRA_DIARIO)}   |   "
        f"Entregadas este ciclo: {entregadas}   |   "
        f"Alertas +72h: {alertas}"
    )
    c.font = Font(name="Calibri", size=9, color="555555")
    c.fill = _fill(TITLE_GRAY)
    c.alignment = _left()
    ws.row_dimensions[2].height = 18

    # ── Fila 3: espacio ───────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Fila 4: encabezados ────────────────────────────────────────────────────
    HEADERS = [
        "Fecha",
        "Hora",
        "# Guía",
        "Asesor",
        "Cliente",
        "Estado",
        "Estado Raw (TCC)",
        "Última Novedad",
        "Horas s/Mov.",
        "Días s/Mov.",
        "Entregado",
        "Alerta +72h",
        "Observaciones",
    ]

    HDR_ROW = 4
    for col, header in enumerate(HEADERS, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=header)
        c.font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill  = _fill(NAVY)
        c.border = _border()
        c.alignment = _center()
    ws.row_dimensions[HDR_ROW].height = 26

    # Autofilter
    ws.auto_filter.ref = f"A{HDR_ROW}:M{HDR_ROW + len(MUESTRA_DIARIO)}"
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    # ── Filas de datos ─────────────────────────────────────────────────────────
    for i, row in enumerate(MUESTRA_DIARIO, 1):
        guia, asesor, cliente, estado, raw, novedad, horas, entregado, alerta, obs = row
        excel_row = HDR_ROW + i

        # Color de fondo según condición
        if entregado:
            bg = DELIVERED
        elif alerta:
            bg = ALERT
        elif estado == "Novedad":
            bg = NOVEDAD
        elif i % 2 == 0:
            bg = ROW_ALT
        else:
            bg = ROW_LIGHT

        bold = entregado or alerta

        valores = [
            HOY,
            HORA,
            guia,
            asesor,
            cliente,
            estado,
            raw,
            novedad if novedad else "—",
            f"{horas:.1f}" if horas > 0 else "0.0",
            f"{horas/24:.2f}" if horas > 0 else "0.00",
            "✓ Sí" if entregado else "No",
            "⚠ Sí" if alerta    else "No",
            obs if obs else "",
        ]

        for col, val in enumerate(valores, 1):
            c = ws.cell(row=excel_row, column=col, value=val)
            c.font   = _font(bold=bold, size=9)
            c.fill   = _fill(bg)
            c.border = _border()
            # Centrado para columnas cortas; izquierda para texto largo
            if col in (1, 2, 9, 10, 11, 12):
                c.alignment = _center()
            else:
                c.alignment = _left()

        ws.row_dimensions[excel_row].height = 20

    # ── Anchos de columna (ajustados a contenido, sin wrap) ───────────────────
    anchos = {
        "A": 14,   # Fecha
        "B": 8,    # Hora
        "C": 18,   # # Guía
        "D": 18,   # Asesor
        "E": 26,   # Cliente
        "F": 16,   # Estado
        "G": 40,   # Estado Raw — columna más ancha para texto TCC
        "H": 20,   # Última Novedad
        "I": 12,   # Horas
        "J": 12,   # Días
        "K": 11,   # Entregado
        "L": 11,   # Alerta
        "M": 38,   # Observaciones
    }
    for col_letter, width in anchos.items():
        ws.column_dimensions[col_letter].width = width

    # ── Leyenda de colores (debajo de la tabla) ────────────────────────────────
    leyenda_row = HDR_ROW + len(MUESTRA_DIARIO) + 2

    leyendas = [
        (DELIVERED, "Verde: Entregado en este ciclo"),
        (ALERT,     "Rojo claro: Alerta +72h sin movimiento"),
        (NOVEDAD,   "Amarillo: Novedad / problema registrado"),
    ]
    ws.cell(row=leyenda_row, column=1, value="Leyenda:").font = _font(bold=True, size=9)

    for j, (color, texto) in enumerate(leyendas, 1):
        col_inicio = 2 + (j - 1) * 3
        c_color = ws.cell(row=leyenda_row, column=col_inicio)
        c_color.fill  = _fill(color)
        c_color.border = _border()
        ws.column_dimensions[get_column_letter(col_inicio)].width = max(
            ws.column_dimensions[get_column_letter(col_inicio)].width, 3
        )
        ws.merge_cells(
            start_row=leyenda_row, start_column=col_inicio + 1,
            end_row=leyenda_row,   end_column=col_inicio + 2
        )
        c_txt = ws.cell(row=leyenda_row, column=col_inicio + 1, value=texto)
        c_txt.font = _font(size=8, color="444444")
        c_txt.alignment = _left()

    ws.row_dimensions[leyenda_row].height = 18


def generar_consolidado_semanal(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Consolidado Semanal")

    SEMANA = "2026-04-13 al 2026-04-18"

    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "ASTECO — Consolidado Semanal de Guías TCC"
    c.font  = Font(name="Calibri", bold=True, size=14, color=NAVY)
    c.fill  = _fill(TITLE_GRAY)
    c.alignment = _left()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    entregadas = sum(1 for r in MUESTRA_SEMANAL if r[5])
    activas    = sum(1 for r in MUESTRA_SEMANAL if r[7])
    alertas    = sum(1 for r in MUESTRA_SEMANAL if r[8])
    c.value = (
        f"Semana: {SEMANA}   |   "
        f"Total guías: {len(MUESTRA_SEMANAL)}   |   "
        f"Entregadas: {entregadas}   |   "
        f"Activas al cierre: {activas}   |   "
        f"Con alertas: {alertas}"
    )
    c.font = Font(name="Calibri", size=9, color="555555")
    c.fill = _fill(TITLE_GRAY)
    c.alignment = _left()
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    HEADERS = [
        "Semana",
        "# Guía",
        "Asesor",
        "Cliente",
        "Primer Estado",
        "Último Estado",
        "Fecha Entrega",
        "Movimientos",
        "Activa al Cierre",
        "Alertas",
        "Observaciones",
    ]
    HDR_ROW = 4
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill  = _fill(NAVY)
        c.border = _border()
        c.alignment = _center()
    ws.row_dimensions[HDR_ROW].height = 26

    ws.auto_filter.ref = f"A{HDR_ROW}:K{HDR_ROW + len(MUESTRA_SEMANAL)}"
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    for i, row in enumerate(MUESTRA_SEMANAL, 1):
        guia, asesor, cliente, prim, ult, entrega, movs, activa, nalert, obs = row
        excel_row = HDR_ROW + i

        if entrega:
            bg = DELIVERED
        elif nalert > 0:
            bg = ALERT
        elif i % 2 == 0:
            bg = ROW_ALT
        else:
            bg = ROW_LIGHT

        valores = [
            SEMANA,
            guia,
            asesor,
            cliente,
            prim,
            ult,
            entrega if entrega else "—",
            movs,
            "Sí" if activa else "No",
            str(nalert) if nalert else "—",
            obs if obs else "",
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=excel_row, column=col, value=val)
            c.font   = _font(size=9)
            c.fill   = _fill(bg)
            c.border = _border()
            if col in (7, 8, 9, 10):
                c.alignment = _center()
            else:
                c.alignment = _left()
        ws.row_dimensions[excel_row].height = 20

    anchos_sem = {
        "A": 26, "B": 18, "C": 18, "D": 26,
        "E": 18, "F": 18, "G": 20, "H": 13,
        "I": 16, "J": 10, "K": 35,
    }
    for col_letter, width in anchos_sem.items():
        ws.column_dimensions[col_letter].width = width


# ── Main ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    out_dir = Path(__file__).parent.parent / "reports" / "muestra"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "reporte_tcc_MUESTRA_2026-04-22_0700.xlsx"

    wb = openpyxl.Workbook()
    generar_reporte_diario(wb)
    generar_consolidado_semanal(wb)
    wb.save(str(out_path))

    print(f"\n✓ Excel generado exitosamente:")
    print(f"  {out_path}")
    print(f"\n  Hoja 1: Reporte Diario — {len(MUESTRA_DIARIO)} guías")
    print(f"  Hoja 2: Consolidado Semanal — {len(MUESTRA_SEMANAL)} guías")
    print(f"\n  Tamaño: {out_path.stat().st_size:,} bytes")
