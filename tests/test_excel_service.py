"""Tests de generación Excel — verifica que los archivos se crean y tienen contenido válido."""

import tempfile
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from app.services.excel_service import DailyReportRow, ExcelService, WeeklyReportRow

_svc = ExcelService()


def _make_daily_row(
    tracking: str = "TCC-001",
    delivered: bool = False,
    is_alert: bool = False,
) -> DailyReportRow:
    return DailyReportRow(
        query_date=date(2026, 4, 22),
        query_time="07:00",
        tracking_number=tracking,
        advisor_name="Asesor Test",
        client_name="Cliente Test",
        current_status="en_transito",
        current_status_raw="EN TRÁNSITO - PLANTA BOGOTÁ",
        last_event_at=datetime(2026, 4, 20, 10, 0, tzinfo=timezone.utc),
        hours_without_movement=48.5,
        days_without_movement=2.02,
        is_delivered=delivered,
        is_alert=is_alert,
        observations="Test",
    )


def _make_weekly_row(tracking: str = "TCC-001") -> WeeklyReportRow:
    return WeeklyReportRow(
        week_label="2026-04-13 al 2026-04-18",
        tracking_number=tracking,
        advisor_name="Asesor Test",
        client_name="Cliente Test",
        first_status="recogido",
        last_status="entregado",
        delivered_at=datetime(2026, 4, 17, 14, 0, tzinfo=timezone.utc),
        total_movements=5,
        still_active=False,
        alerts_detected=0,
        observations="",
    )


def test_generate_daily_creates_file():
    rows = [_make_daily_row("TCC-001"), _make_daily_row("TCC-002", is_alert=True)]
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "test_daily.xlsx"
        result = _svc.generate_daily(rows, path, "0700", date(2026, 4, 22))
        assert result.exists()
        assert result.stat().st_size > 0


def test_daily_excel_has_correct_headers():
    rows = [_make_daily_row()]
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "daily.xlsx"
        _svc.generate_daily(rows, path, "0700", date(2026, 4, 22))
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Encabezados están en fila 4 (3 filas de título + 1 header)
        header_row = [ws.cell(row=4, column=c).value for c in range(1, 14)]
        assert "# Guía" in header_row
        assert "Asesor" in header_row
        assert "Estado Actual" in header_row
        assert "Alerta 72h" in header_row


def test_daily_excel_data_rows():
    rows = [_make_daily_row("TCC-999")]
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "daily.xlsx"
        _svc.generate_daily(rows, path, "1200", date(2026, 4, 22))
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        # Primera fila de datos es la fila 5
        guia_cell = ws.cell(row=5, column=3).value  # columna 3 = # Guía
        assert guia_cell == "TCC-999"


def test_daily_excel_delivered_row():
    rows = [_make_daily_row("TCC-DELIVERED", delivered=True)]
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "daily.xlsx"
        _svc.generate_daily(rows, path, "1600", date(2026, 4, 22))
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        entregado_cell = ws.cell(row=5, column=11).value  # columna 11 = Entregado
        assert entregado_cell == "Sí"


def test_generate_weekly_creates_file():
    rows = [_make_weekly_row("TCC-001"), _make_weekly_row("TCC-002")]
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "weekly.xlsx"
        result = _svc.generate_weekly(rows, date(2026, 4, 13), date(2026, 4, 18), path)
        assert result.exists()
        assert result.stat().st_size > 0


def test_generate_weekly_headers():
    rows = [_make_weekly_row()]
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "weekly.xlsx"
        _svc.generate_weekly(rows, date(2026, 4, 13), date(2026, 4, 18), path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        header_row = [ws.cell(row=4, column=c).value for c in range(1, 12)]
        assert "# Guía" in header_row
        assert "Total Movimientos" in header_row
        assert "Siguió Activa" in header_row


def test_daily_empty_rows_still_creates_file():
    """Con lista vacía de rows debe crear archivo sin error."""
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "empty.xlsx"
        _svc.generate_daily([], path, "0700", date(2026, 4, 22))
        assert path.exists()


def test_generate_creates_parent_dirs():
    """El servicio debe crear directorios padre si no existen."""
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "sub1" / "sub2" / "report.xlsx"
        _svc.generate_daily([_make_daily_row()], path, "0700", date(2026, 4, 22))
        assert path.exists()
